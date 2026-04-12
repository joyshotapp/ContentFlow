"""Excel 匯入工具 — 將 SEO 專案管理表匯入資料庫"""

import re
from pathlib import Path

import openpyxl
from loguru import logger
from sqlalchemy.orm import Session

from contentflow.db import get_db, init_db
from contentflow.models.database import (
    Article,
    Category,
    CategorySEO,
    Changelog,
    Competitor,
    ContentCalendar,
    ContentStrategy,
    Keyword,
    LegalTerm,
    Product,
    SEORanking,
    WritingRule,
)


PROJECT_SCOPED_MODELS = (
    Keyword,
    Category,
    ContentCalendar,
    Article,
    WritingRule,
    ContentStrategy,
    Competitor,
    Product,
    LegalTerm,
    SEORanking,
    CategorySEO,
    Changelog,
)


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _clean_keyword_text(value) -> str:
    text = _safe_str(value)
    text = re.sub(r'\s*[\(（]\d+[\)）]\s*', ' ', text)
    text = re.sub(r'\s+\d+$', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _split_keyword_lines(value) -> list[str]:
    items = []
    seen = set()
    for part in re.split(r'[\n,，]+', _safe_str(value)):
        cleaned = _clean_keyword_text(part)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        items.append(cleaned)
    return items


def _safe_int(value) -> int:
    if value is None:
        return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def import_keywords(ws, session: Session, project_id: int | None = None):
    """匯入關鍵字表"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    count = 0
    for row in rows[1:]:  # 跳過標題列
        if all(v is None for v in row):
            continue
        keyword_text = _safe_str(row[2]) if len(row) > 2 else ""
        if not keyword_text:
            continue

        kw = Keyword(
            priority=_safe_str(row[0]),
            usage=_safe_str(row[1]),
            keyword=keyword_text,
            search_volume=_safe_float(row[3]) if len(row) > 3 else 0,
            cpc=_safe_float(row[4]) if len(row) > 4 else 0,
            paid_difficulty=_safe_float(row[5]) if len(row) > 5 else 0,
            seo_difficulty=_safe_float(row[6]) if len(row) > 6 else 0,
            steve_note=_safe_str(row[8]) if len(row) > 8 else "",
            project_id=project_id,
        )
        session.add(kw)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆關鍵字")
    return count


def import_categories(ws_slug, session: Session, project_id: int | None = None):
    """從分類slug sheet 匯入分類"""
    rows = list(ws_slug.iter_rows(values_only=True))
    count = 0

    for row in rows[1:]:  # 跳過標題
        if all(v is None for v in row):
            continue
        name = _safe_str(row[0])
        slug = _safe_str(row[1]) if len(row) > 1 else ""
        if not name or not slug:
            continue

        cat_type_raw = _safe_str(row[2]) if len(row) > 2 else ""
        cat_type = cat_type_raw if cat_type_raw in {"category", "tag"} else "category"
        cat = Category(
            name=name,
            slug=slug,
            cat_type=cat_type,
            project_id=project_id,
        )
        session.add(cat)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆分類/標籤")
    return count


def import_content_calendar(ws, session: Session, project_id: int | None = None):
    """匯入 2026 執行內容（即(ing)文章規劃 sheet 的月度排程）"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    for row in rows:
        if all(v is None for v in row):
            continue
        # 根據 Excel 結構：col5=月, col6=週, col7=類型, col8=標題, col9=關鍵字,
        # col10=搜尋意圖, col11=受眾, col12=架構, col13=FAQ
        month = _safe_int(row[5]) if len(row) > 5 else 0
        week = _safe_int(row[6]) if len(row) > 6 else 0
        if month == 0 or week == 0:
            continue

        title = _safe_str(row[8]) if len(row) > 8 else ""
        if not title:
            continue

        cal = ContentCalendar(
            month=month,
            week=week,
            article_type=_safe_str(row[7]) if len(row) > 7 else "",
            title=title,
            keywords=_safe_str(row[9]) if len(row) > 9 else "",
            search_intent=_safe_str(row[10]) if len(row) > 10 else "",
            target_audience=_safe_str(row[11]) if len(row) > 11 else "",
            writing_architecture=_safe_str(row[12]) if len(row) > 12 else "",
            faq_questions=_safe_str(row[13]) if len(row) > 13 else "",
            project_id=project_id,
        )
        session.add(cal)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆內容日曆")
    return count


def import_articles(ws, session: Session, project_id: int | None = None):
    """匯入文章規劃 (ing)文章規劃"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    count = 0

    # 先掃出所有明確 seqno，用於自動編號
    explicit_seqnos = set()
    for row in rows[1:]:  # 跳過標題列
        if all(v is None for v in row):
            continue
        s = _safe_int(row[0]) if len(row) > 0 else 0
        if s > 0:
            explicit_seqnos.add(s)
    next_auto = (max(explicit_seqnos) + 1) if explicit_seqnos else 1

    for row in rows[1:]:  # 跳過標題列
        if all(v is None for v in row):
            continue
        seqno = _safe_int(row[0]) if len(row) > 0 else 0
        primary_kw = _safe_str(row[1]) if len(row) > 1 else ""
        # 有些行只有 seqno 為 None 但有其他欄位，也有些行只有待選關鍵字
        if not primary_kw and seqno == 0:
            continue

        # 自動編號
        if seqno <= 0:
            seqno = next_auto
            next_auto += 1

        primary_parts = _split_keyword_lines(primary_kw)
        secondary_parts = _split_keyword_lines(row[2]) if len(row) > 2 else []
        merged_secondary = []
        if primary_parts:
            kw_clean = primary_parts[0]
            seen = {kw_clean}
            for keyword in primary_parts[1:] + secondary_parts:
                if keyword not in seen:
                    seen.add(keyword)
                    merged_secondary.append(keyword)
        else:
            kw_clean = ""
        # 解析主關鍵字搜索量，只取第一行/第一組關鍵字的量
        first_line = _safe_str(primary_kw).splitlines()[0] if _safe_str(primary_kw) else ""
        vol_match = re.search(r'(?:\((\d+)\)|(\d+)\s*$)', first_line)
        volume = float(vol_match.group(1) or vol_match.group(2)) if vol_match else 0

        article = Article(
            seqno=seqno,
            primary_keyword=kw_clean,
            primary_keyword_volume=volume,
            secondary_keywords="\n".join(merged_secondary),
            outline=_safe_str(row[3]) if len(row) > 3 else "",
            google_doc_url=_safe_str(row[4]) if len(row) > 4 else "",
            title=_extract_title(row[3]) if len(row) > 3 else "",
            status="planned",
            project_id=project_id,
        )
        session.add(article)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆文章規劃")
    return count


def _extract_title(outline_text) -> str:
    """從文章架構中提取建議標題"""
    if not outline_text:
        return ""
    text = str(outline_text)
    match = re.search(r'建議標題[:：]\s*(.+?)[\n\r]', text)
    if match:
        return match.group(1).strip()
    return ""


def import_writing_rules(ws, session: Session, project_id: int | None = None):
    """匯入撰寫規範"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    for row in rows:
        if all(v is None for v in row):
            continue
        content = _safe_str(row[0])
        if not content:
            continue

        # 解析四種架構
        architectures = [
            ("倒三角架構", "architecture"),
            ("金字塔架構", "architecture"),
            ("思維流程型", "architecture"),
            ("敘事型", "architecture"),
        ]

        # 整塊存為一條規則
        rule = WritingRule(
            rule_type="architecture_guide",
            name="四大文章架構運用指南",
            content=content,
            order_num=count,
            project_id=project_id,
        )
        session.add(rule)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆撰寫規範")
    return count


def import_content_strategy(ws, session: Session, project_id: int | None = None):
    """匯入部落格內容定位"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    order = 0

    for row in rows:
        if all(v is None for v in row):
            continue

        section = _safe_str(row[0])
        content = _safe_str(row[1]) if len(row) > 1 else ""
        extra1 = _safe_str(row[2]) if len(row) > 2 else ""
        extra2 = _safe_str(row[3]) if len(row) > 3 else ""

        # 合併所有非空欄位
        full_content = "\n".join(filter(None, [content, extra1, extra2]))
        if not section and not full_content:
            continue

        cs = ContentStrategy(
            section=section if section else "supplement",
            title=section,
            content=full_content,
            order_num=order,
            project_id=project_id,
        )
        session.add(cs)
        count += 1
        order += 1

    session.flush()
    logger.info(f"匯入 {count} 筆內容策略")
    return count


def import_competitors(ws, session: Session, project_id: int | None = None):
    """匯入競業市場研究"""
    rows = list(ws.iter_rows(values_only=True))

    # 找到「網站」行作為品牌名稱來源
    brand_names = []
    brand_urls = []
    features_list = []
    analysis_list = []
    sells_list = []
    recommendation_list = []

    for row in rows:
        if all(v is None for v in row):
            continue
        first = _safe_str(row[0]) if row else ""

        if first == "網站":
            brand_names = [_safe_str(row[i]) if len(row) > i else "" for i in range(1, 8)]
        elif brand_names and not brand_urls:
            # URL 行（緊跟在品牌名稱後面的空第一欄行）
            if not first:
                brand_urls = [_safe_str(row[i]) if len(row) > i else "" for i in range(1, 8)]
        elif first == "特色":
            features_list = [_safe_str(row[i]) if len(row) > i else "" for i in range(1, 8)]
        elif first == "網站內容經營分析":
            analysis_list = [_safe_str(row[i]) if len(row) > i else "" for i in range(1, 8)]
        elif first == "販售產品":
            sells_list = [_safe_str(row[i]) if len(row) > i else "" for i in range(1, 8)]

    # 找推薦行
    for row in rows:
        if all(v is None for v in row):
            continue
        first_cells = [_safe_str(row[i]) if len(row) > i else "" for i in range(2)]
        if any("模仿" in c for c in first_cells):
            recommendation_list = [_safe_str(row[i]) if len(row) > i else "" for i in range(1, 8)]
            break

    count = 0
    for i, name in enumerate(brand_names):
        if not name:
            continue
        comp = Competitor(
            brand_name=name,
            website=brand_urls[i] if i < len(brand_urls) else "",
            features=features_list[i] if i < len(features_list) else "",
            content_analysis=analysis_list[i] if i < len(analysis_list) else "",
            sells_products=sells_list[i] if i < len(sells_list) else "",
            recommendation=recommendation_list[i] if i < len(recommendation_list) else "",
            project_id=project_id,
        )
        session.add(comp)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆競品資料")
    return count


def import_products(ws, session: Session, project_id: int | None = None):
    """匯入產品資訊"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    # 產品資訊比較精簡，直接存每行
    for row in rows:
        if all(v is None for v in row):
            continue
        content = _safe_str(row[0])
        if not content:
            continue
        prod = Product(
            series_name=content,
            description=content,
            project_id=project_id,
        )
        session.add(prod)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆產品資訊")
    return count


def import_legal_terms(ws, session: Session, project_id: int | None = None):
    """匯入食品廣告用詞規定"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    for row in rows:
        if all(v is None for v in row):
            continue
        content = _safe_str(row[0])
        source = _safe_str(row[1]) if len(row) > 1 else ""
        if not content:
            continue

        # 分類判斷
        if "醫療效能" in content:
            term_type = "forbidden"
            category = "醫療效能"
        elif "誇張" in content or "易生誤解" in content:
            term_type = "caution"
            category = "誇張易誤解"
        elif "可使用" in content or "可敘述" in content or "通常" in content:
            term_type = "allowed"
            category = "可使用"
        else:
            term_type = "reference"
            category = "參考"

        lt = LegalTerm(
            term_type=term_type,
            category=category,
            content=content,
            source=source,
            project_id=project_id,
        )
        session.add(lt)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆法規用詞")
    return count


def import_seo_rankings(ws, session: Session, project_id: int | None = None):
    """匯入 SEO 關鍵字排名表"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    # 找到排名日期和搜尋引擎
    tracked_date = ""
    search_engine = "Google"
    for row in rows:
        if len(row) > 2:
            if _safe_str(row[1]) == "排名日期":
                tracked_date = _safe_str(row[2])
            elif _safe_str(row[1]) == "搜尋引擎":
                search_engine = _safe_str(row[2])
            elif _safe_str(row[1]) == "關鍵字":
                continue  # 標題行
            elif _safe_int(row[0]) > 0:
                kw = _safe_str(row[1])
                if kw:
                    ranking = SEORanking(
                        keyword=kw,
                        rank=_safe_int(row[2]) if len(row) > 2 else None,
                        landing_page=_safe_str(row[3]) if len(row) > 3 else "",
                        search_engine=search_engine,
                        tracked_date=tracked_date,
                        project_id=project_id,
                    )
                    session.add(ranking)
                    count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆排名資料")
    return count


def import_category_seo(ws, session: Session, project_id: int | None = None):
    """匯入分類規劃、關鍵字配置"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    for row in rows[1:]:  # 跳過標題
        if all(v is None for v in row):
            continue
        level2 = _safe_str(row[1]) if len(row) > 1 else ""
        level3 = _safe_str(row[2]) if len(row) > 2 else ""
        if not level2 and not level3:
            continue

        cs = CategorySEO(
            level2=level2,
            level3=level3,
            meta_keywords=_safe_str(row[3]) if len(row) > 3 else "",
            original_title=_safe_str(row[4]) if len(row) > 4 else "",
            original_description=_safe_str(row[5]) if len(row) > 5 else "",
            new_title=_safe_str(row[6]) if len(row) > 6 else "",
            new_description=_safe_str(row[7]) if len(row) > 7 else "",
            notes=_safe_str(row[8]) if len(row) > 8 else "",
            project_id=project_id,
        )
        session.add(cs)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆分類 SEO 配置")
    return count


def import_changelog(ws, session: Session, project_id: int | None = None):
    """匯入 Shopify code change-log"""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    for row in rows[1:]:  # 跳過標題
        if all(v is None for v in row):
            continue
        filename = _safe_str(row[1]) if len(row) > 1 else ""
        if not filename:
            continue

        cl = Changelog(
            theme_version=_safe_str(row[0]),
            filename=filename,
            original_code=_safe_str(row[2]) if len(row) > 2 else "",
            new_code=_safe_str(row[3]) if len(row) > 3 else "",
            project_id=project_id,
        )
        session.add(cl)
        count += 1

    session.flush()
    logger.info(f"匯入 {count} 筆 changelog")
    return count


def _clear_existing_data(session: Session, project_id: int | None = None) -> None:
    """清除既有資料，但不刪除 projects 表。"""
    for model in PROJECT_SCOPED_MODELS:
        query = session.query(model)
        if project_id is not None:
            query = query.filter(model.project_id == project_id)
        query.delete(synchronize_session=False)
    session.commit()


# ── 主匯入函数 ────────────────────────────────────────────────

SHEET_MAP = {
    "關鍵字表": import_keywords,
    "分類slug": import_categories,
    "(ing)文章規劃": import_articles,
    "撰寫規範": import_writing_rules,
    "部落格內容定位": import_content_strategy,
    "競業市場研究": import_competitors,
    "產品資訊": import_products,
    "食品廣告用詞規定": import_legal_terms,
    "SEO關鍵字排名表": import_seo_rankings,
    "分類規劃、關鍵字配置": import_category_seo,
    "shopify code change-log": import_changelog,
}

# 2026執行內容 的內容日曆需特別處理 — 它在「(ing)文章規劃」同一 sheet
CALENDAR_SHEET = "2026執行內容&成效重點"


def import_excel(excel_path: str, clear_existing: bool = True, project_id: int | None = None) -> dict:
    """
    完整匯入 Excel 到資料庫

    Args:
        excel_path: Excel 檔案路徑
        clear_existing: 是否清除現有資料（預設 True，全量匯入）
        project_id: 指定專案 ID（None 代表不指定/沿用既有關聯）

    Returns:
        各表匯入數量的字典
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 Excel 檔案: {excel_path}")

    logger.info(f"開始匯入 Excel: {path.name}")
    wb = openpyxl.load_workbook(str(path))
    logger.info(f"偵測到 {len(wb.sheetnames)} 個工作表: {wb.sheetnames}")

    # 初始化資料庫
    init_db()
    session = get_db()
    results = {}

    try:
        # 清除現有資料
        if clear_existing:
            _clear_existing_data(session, project_id=project_id)
            if project_id is not None:
                logger.info(f"已清除專案 {project_id} 的既有資料")
            else:
                logger.info("已清除所有專案內容資料（保留 projects 表）")

        # 匯入各 sheet
        for sheet_name, importer in SHEET_MAP.items():
            # 模糊匹配 sheet 名（有些名稱可能含空格或不同字元）
            matched = None
            for ws_name in wb.sheetnames:
                if sheet_name in ws_name or ws_name in sheet_name:
                    matched = ws_name
                    break

            if matched:
                ws = wb[matched]
                try:
                    count = importer(ws, session, project_id=project_id)
                    results[sheet_name] = count
                except Exception as e:
                    logger.error(f"匯入 [{sheet_name}] 失敗: {e}")
                    results[sheet_name] = f"ERROR: {e}"
            else:
                logger.warning(f"找不到工作表: {sheet_name}")
                results[sheet_name] = 0

        # 匯入 2026 內容日曆（從「2026年平台內容主題參考」sheet）
        cal_matched = None
        for ws_name in wb.sheetnames:
            if "2026" in ws_name and "主題" in ws_name:
                cal_matched = ws_name
                break
        # fallback
        if not cal_matched:
            for ws_name in wb.sheetnames:
                if "2026" in ws_name and "內容" in ws_name and "執行" not in ws_name:
                    cal_matched = ws_name
                    break

        if cal_matched:
            try:
                count = import_content_calendar(wb[cal_matched], session, project_id=project_id)
                results["內容日曆"] = count
            except Exception as e:
                logger.error(f"匯入 [內容日曆] 失敗: {e}")
                results["內容日曆"] = f"ERROR: {e}"

        session.commit()
        logger.info(f"匯入完成！總結: {results}")
        return results

    except Exception as e:
        session.rollback()
        logger.error(f"匯入過程中發生錯誤: {e}")
        raise
    finally:
        session.close()
        wb.close()


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python -m contentflow.tools.excel_importer <excel_path>")
        sys.exit(1)

    results = import_excel(sys.argv[1])
    print("\n=== 匯入結果 ===")
    for sheet, count in results.items():
        print(f"  {sheet}: {count}")
